#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
检查QTP检查单中的规则是否都已写入DO160G规则文件
"""
import openpyxl
import os
import sys
from collections import defaultdict

# 设置输出编码
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')

# 读取Excel文件
excel_path = "./prompts/QTP检查单-全部规则-PDDS.xlsx"
rules_dir = "./prompts/DO160G规则/"

print("=" * 80)
print("Step 1: Reading QTP checklist Excel file...")
print("=" * 80)

wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# 先查看Excel的列结构
print(f"\nExcel sheet name: {ws.title}")
print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")

# 打印表头
print("\nHeader row:")
header_row = []
for idx, cell in enumerate(ws[1], 1):
    value = cell.value
    header_row.append(value)
    print(f"  Column {idx}: {value}")

# 收集所有数据
print("\n" + "=" * 80)
print("Step 2: Extracting all rules from Excel...")
print("=" * 80)

all_rows = []
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    if any(cell for cell in row):  # 至少有一个非空单元格
        row_data = {
            'row_num': row_idx,
            'cells': [str(cell) if cell is not None else "" for cell in row]
        }
        all_rows.append(row_data)

print(f"\nTotal data rows: {len(all_rows)}")
print("\nFirst 10 rows sample:")
for row_data in all_rows[:10]:
    cells = row_data['cells']
    print(f"Row {row_data['row_num']}: {cells[:3]}...")  # 只显示前3列

# 读取规则文件
print("\n" + "=" * 80)
print("Step 3: Reading DO160G rule files...")
print("=" * 80)

rule_files = {}
for filename in sorted(os.listdir(rules_dir)):
    if filename.endswith('.md'):
        filepath = os.path.join(rules_dir, filename)
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
            rule_files[filename] = content
            print(f"  [{len(content):7d} bytes] {filename}")

print(f"\nTotal rule files: {len(rule_files)}")

# 输出到文件
output_path = "./rule_comparison_result.txt"
with open(output_path, 'w', encoding='utf-8') as f:
    f.write("QTP检查单规则对比分析结果\n")
    f.write("=" * 80 + "\n\n")

    f.write("一、Excel表头结构\n")
    f.write("-" * 80 + "\n")
    for idx, col_name in enumerate(header_row, 1):
        f.write(f"列 {idx}: {col_name}\n")

    f.write(f"\n二、Excel数据行数: {len(all_rows)}\n")
    f.write("-" * 80 + "\n")
    f.write("前20行数据示例:\n\n")
    for row_data in all_rows[:20]:
        f.write(f"第{row_data['row_num']}行: {row_data['cells']}\n")

    f.write(f"\n三、DO160G规则文件列表 (共{len(rule_files)}个)\n")
    f.write("-" * 80 + "\n")
    for filename in sorted(rule_files.keys()):
        f.write(f"  {filename}\n")

print(f"\nAnalysis result saved to: {output_path}")
print("\n" + "=" * 80)
print("Analysis complete!")
print("=" * 80)
